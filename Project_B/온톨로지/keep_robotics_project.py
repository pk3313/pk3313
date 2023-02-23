#-*- coding:utf-8 -*-

import urllib3
import json
import csv
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

##---------- init ----------##
count = True
#일반 명사, 고유명사를 추출하기 위한 태그명
des=["NNG","NNP"]
fileName = []
fileNum = "0000"
path = 'C:/Users/work/Desktop/keep_robotics_project/2nd_meet/code/동화' 

#.txt 제거하고 파일 이름만 저장
file_list = os.listdir(path)
for i in file_list:
    f =i.split('.')
    fileName.append(f[0])
  
loopCount = 80
i = 0
j = 0
describe = []
verb=[]
wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1.title = "Video Transcripts"
sheet2 = wb.create_sheet("Sentences")
sheet3 = wb.create_sheet("words")
sheet4 = wb.create_sheet("Verb")

##---------- ETRI API ----------##
openApiURL = "http://aiopen.etri.re.kr:8000/WiseNLU" # 언어 분석 기술(문어)
openApiURL = "http://aiopen.etri.re.kr:8000/WiseNLU_spoken" # 언어 분석 기술(구어)
accessKey = "31e6f9bb-3afc-48ba-90d8-fd988cf15f2e"
analysisCode = "ner"
analysisCode_verb= "morp"
analysisCode_sentence= "dparse"
text = ""

# 중복 제거_동사
def dropDuplicates_desc(data):
    f = open('word_final_describe.csv', 'w',newline='',encoding='cp949')
    csv_f = csv.writer(f)
    word =[]
    for i in data:    
        cnt = i.split(',')
        if "Name" not in cnt or "Count\n" not in cnt:
            if len(cnt) ==2:
                num = int(cnt[1])
                if num >=2:
                    wo = i.split(',')
                    word.append(wo[0])
            else :
                wo = i.split(',')
                word.append(wo[0])     
    word =list(set(word))
    for i in word:
        csv_f.writerow([i])    
    f.close()    
   
def dropDuplicates_verb(data):#중복 제거_명사 
    f = open('word_final_verb.csv', 'w',newline='',encoding='cp949')
    csv_f = csv.writer(f)
    word =[]
    for i in data:    
        cnt = i.split(',')
        if "Name" not in cnt or "Count\n" not in cnt:
            if len(cnt) ==2:
                num = int(cnt[1])
                if num >=2:
                    wo = i.split(',')
                    word.append(wo[0])
            else :
                wo = i.split(',')
                word.append(wo[0])           
    word =list(set(word))
    for i in word:
        csv_f.writerow([i])            
    f.close()

#최종 중복 제거
def final():
    sheet3 = pd.read_excel('sentence_test.xlsx',engine='openpyxl',sheet_name='words')
    sheet4 = pd.read_excel('sentence_test.xlsx',engine='openpyxl',sheet_name='Verb')
    df_word = sheet3.drop_duplicates( keep='first')
    df_verb = sheet4.drop_duplicates( keep='first')
    with pd.ExcelWriter('sentence_test.xlsx', engine='openpyxl',mode ='a')as wr:
        df_word.to_excel(wr,sheet_name='Words_final', index=False ,columns=None, header=None)
        df_verb.to_excel(wr,sheet_name='Verb_final', index=False  ,columns=None, header=None)
    wb = openpyxl.load_workbook('sentence_test.xlsx')     
    wb.remove_sheet(wb['words'])
    wb.remove_sheet(wb['Verb'])
    wb.save('sentence_test.xlsx')
    wb.close()  
def Verb_def():
    requestJson = {
        "access_key": accessKey,
        "argument": {
            "text": text,
            "analysis_code": analysisCode_verb}}
    http = urllib3.PoolManager()
    response = http.request(
        "POST",
        openApiURL,
        headers={"Content-Type": "application/json; charset=UTF-8"},
        body=json.dumps(requestJson))
    responseToJson = json.loads(str(response.data,"utf-8"))
    for i in range(len(responseToJson["return_object"]["sentence"])):
        for j in range(len(responseToJson["return_object"]["sentence"][i]["morp"])):
            if responseToJson["return_object"]["sentence"][i]["morp"][j]["type"] == "VV":
                verb.append(responseToJson["return_object"]["sentence"][i]["morp"][j]["lemma"])
def requestFunc():
    requestJson = {
        "access_key": accessKey,
        "argument": {
            "text": text,
            "analysis_code": analysisCode}}
    http = urllib3.PoolManager()
    response = http.request(
        "POST",
        openApiURL,
        headers={"Content-Type": "application/json; charset=UTF-8"},
        body=json.dumps(requestJson))
    responseToJson = json.loads(str(response.data,"utf-8"))
    for i in range(len(responseToJson["return_object"]["sentence"])):
         for j in range(len(responseToJson["return_object"]["sentence"][i]["morp"])):
            if responseToJson["return_object"]["sentence"][i]["morp"][j]["type"] in des:
                describe.append(responseToJson["return_object"]["sentence"][i]["morp"][j]["lemma"])
def Story(): # excel로 저장  
    num =0
    aaa = 1
    bbb = 0
    number =1
    num_word =1
    num_verb =1
    da = '다'
    for i in file_list:
        f= open('C:/Users/work/Desktop/keep_robotics_project/2nd_meet/code/동화/'+i,'r',newline='',encoding='utf-8')
        f_des = open('word_final_describe.csv','r',newline='',encoding='cp949')
        f_verb = open('word_final_verb.csv','r',newline='',encoding='cp949')  
        file = f.readlines()
        tempstr = ""
        for o in file: # 원본 텍스트 저장
            tempstr += o
        sheet1['A'+str(aaa)]= fileName[bbb]
        sheet1['A'+str(aaa)].font = Font(bold=True,size=15)
        bbb += 1
        aaa+=1
        sheet1.column_dimensions['A'].width=300
        sheet1['A'+str(aaa)].font = Font(size=10)
        sheet1['A'+str(aaa)].alignment = Alignment(wrap_text=True)
        sheet1.row_dimensions[aaa].height=400
        sheet1['A'+str(aaa)]= tempstr
        aaa+=1
        sheet1['A'+str(aaa)]= ' '
        aaa+=1
        for ii in file: # 문장으로 쪼개기
            lens = ii.split('.')
            for i in range(len(lens)):
                a = lens[i]
                b = a.strip('\r\n "\r\n')
                if b =='”' or b =='' :
                    pass
                else: 
                    num =str(number)
                    sheet2['A'+num]=b
                    number+=1
        for i in f_des: # 단어 저장
            numb = str(num_word)
            sheet3['A'+numb] = i
            num_word+=1
        for i in f_verb:  # 동사 저장
            numb_verb = str(num_verb)
            sheet4['A'+numb_verb] = i+da
            num_verb+=1 
    wb.save('sentence_test.xlsx')
    f.close()

# 해당 경로의 파일을 읽어와 작업 진행
d =0
for jj in file_list:
    d+=1
    f = open('C:/Users/work/Desktop/keep_robotics_project/2nd_meet/code/동화/'+jj,"r",encoding="utf-8")
    cnt = len(f.readlines())
    ii=0
    f.close()

    with open('C:/Users/work/Desktop/keep_robotics_project/2nd_meet/code/동화/'+jj,"r",encoding="utf-8") as json_file:

        for i in json_file:
            
            ii+=1
            text+= i
            j+=1
           
            if ii == cnt:
                print(f'{d}번째 {jj}파일입니다.')
                requestFunc()
                Verb_def()
                print("중복 단어 제거!")
                dropDuplicates_desc(describe)
                dropDuplicates_verb(verb)
                print('엑셀에 저장하는 중입니다.')
                Story() # 엑셀에 저장하는 함수
                text = ""
                count = False
            elif j == loopCount:
                    j = 0
                    requestFunc()
                    Verb_def()
                    text = ""

wb.close()       
print("저장이 완료되었습니다.")                
f.close()
print("최종 중복 제거 실행!")
final()
print('쿠쿠가 맛있는 밥을 완성하였습니다.')                
            
        

